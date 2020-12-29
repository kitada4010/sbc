"""
引数一覧
1 : 読み込みディレクトリのパス
2 : 書き込みディレクトリのパス
3 : 個体ファイルのパス
4 : ファイルの名前
5 : h
6 : k
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import openpyxl
import math
import statistics
from matplotlib.backends.backend_pdf import PdfPages


def diff_x_i(sig, x_i) :
    return [sig[i]-x_i for i in range(len(sig))]



#ファイルの読み込み                                                                                                                         

def write_data(wb_sheet, sig, column_point, start_row, value_title) :
    wb_sheet.cell(column=column_point, row=1, value=value_title)
    for i, sig_data in enumerate(sig) :
        #print(sig_data)
        wb_sheet.cell(column=column_point, row=i+start_row, value=sig_data)






def inspect(wb, file1, h, k):
    #ファイル1のデータカウント
    sheet_df1 = file1.parse(file1.sheet_names, header=None)
    sheet_names1 = file1.sheet_names

    for i, name in enumerate(sheet_names1):
        sheet_df1[i] = file1.parse(name)
        try :
            sig_ori = (sheet_df1[i]["volt"]).values
        except KeyError :
            print("not max data number")
            break
        S_1 = [(max(diff_x_i(sig_ori[i-k:i], sig_ori[i])) + max(diff_x_i(sig_ori[i+1:i+k+1], sig_ori[i])))/2 for i in range(k, len(sig_ori)-k)]
        sig_mean = sum(S_1)/len(S_1) # m'
        sig_sd = statistics.pstdev(S_1) # s'
        #        S_1_fire = [1 if(S_1[i]-sig_mean > h*sig_sd) else 0 for i in range(len(S_1))]
        S_1_fire_dt = [1 if(S_1[i]-sig_mean > h*sig_sd) & (sig_ori[i+k] < sig_ori[i+k-1]) & (sig_ori[i+k] < sig_ori[i+k+1]) else 0 for i in range(len(S_1))]
        new_sig_diff_len = len(sig_ori) - len(S_1_fire_dt)
        write_start_point = new_sig_diff_len/2 + 1
        write_data(wb[name], S_1_fire_dt, 2, write_start_point, "spike_data")
        
#        sig_fire = [ 1 if(sig_smooth[i-3] > sig_smooth[i-2]) & (sig_smooth[i-2] > sig_smooth[i-1]) & (sig_smooth[i-1] > sig_smooth[i]) \
#                   & (sig_smooth[i] < sig_smooth[i+1]) & (sig_smooth[i+1] < sig_smooth[i+2]) & (sig_smooth[i+2] < sig_smooth[i+3]) else 0 for i in range(3, len(sig_smooth)-3)]

    
#read_path = "/st9/b009vb/01data"
read_path = sys.argv[1]
#save_path = ""
save_path = sys.argv[2]
#individual_path = "/restraint-25000/B39-HR/"
individual_path = sys.argv[3]
#file_name = "B39-restraint-R2.xlsx"
file_name = sys.argv[4]
h = float(sys.argv[5])
k = int(sys.argv[6])


file1 = pd.ExcelFile(read_path + individual_path + file_name)
wb = openpyxl.load_workbook(read_path + individual_path + file_name)


inspect(wb, file1, h, k)
wb.save(file_name)
