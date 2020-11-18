import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import math
import openpyxl
import os
from matplotlib.backends.backend_pdf import PdfPages


#引数一覧
"""
sys.argv[1] : 読み込みファイルパス
sys.argv[2] : 書き込みファイルパス
sys.argv[3] : 個体までのパス
sys.argv[4] : ファイル名
"""


#ファイルの読み込み
#file_name1 = sys.argv[1]
#file_name2 = sys.argv[2]
#file1 = pd.ExcelFile(file_name1+'.xlsx')
#file2 = pd.ExcelFile(file_name2pdf = PdfPages('test2.pdf')




def write_data(wb_sheet, sig, column_point) :
        wb_sheet.cell(column=column_point, row=1, value="volt")
        for i, sig_data in enumerate(sig) :
        #print(sig_data)
                wb_sheet.cell(column=column_point, row=i+2, value=sig_data)

def inspect(rb, wb):
        cannel_start = 7
        cannel_end = 2
        sheet_df1 = rb.parse(rb.sheet_names, header=None)
        #ファイル1のデータカウント
        for i, name in enumerate(rb.sheet_names) :
                sheet_df1[i] = rb.parse(name)
        for i in range(len(rb.sheet_names)-1) :
                sheet_new = wb.create_sheet()
        for i , name in enumerate(wb.sheetnames) :
                try :
                        #print(sheet_df1[i][1][1])
                        start_number = (np.where(sheet_df1[i]['INFORMATION']=="CHANNEL")[0][0]) +cannel_start
                        end_number = (np.where(sheet_df1[i]['INFORMATION']=="CHANNEL")[0][1]) - cannel_end
                        sig_ori = (sheet_df1[i]['INFORMATION'][start_number : end_number]).values
                        #           sig1 = (sheet_df1[i]['Unnamed: 3'][start_number : end_number]).values
                except KeyError :
                        print("not max data number")
                        break
                write_data(wb[name], list(sig_ori), 1)



# ディレクトリ作成
def my_makedirs(path):
        if not os.path.isdir(path):
                os.makedirs(path)


#file1 = pd.ExcelFile(sys.argv[1])





#read_path = "/st9/b009vb/01data"
read_path = sys.argv[1]
#save_path = ""
save_path = sys.argv[2]
#individual_path = "/restraint-25000/B39-HR/"
individual_path = sys.argv[3]
#file_name = "B39-restraint-R2.xlsx"
file_name = sys.argv[4]

#sheet_names1 = file1.sheet_names

rb = pd.ExcelFile(read_path + individual_path + file_name)
#file_kull = open("divergence.txt", "a")
#file_data = open("data-ab.txt", "a")
#count_data = open("count_data.csv", "a")

wb = openpyxl.Workbook()
inspect(rb, wb)
#my_makedirs(save_path + individual_path)
wb.save(file_name)
wb.close()
