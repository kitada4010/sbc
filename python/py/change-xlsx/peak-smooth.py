import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import openpyxl
import math
from matplotlib.backends.backend_pdf import PdfPages



#ファイルの読み込み                                                                                                                         

def write_data(wb_sheet, sig, column_point, start_row, value_title) :
    wb_sheet.cell(column=column_point, row=1, value=value_title)
    for i, sig_data in enumerate(sig) :
        #print(sig_data)
        wb_sheet.cell(column=column_point, row=i+start_row, value=sig_data)






def inspect(wb, file1, peak):
    #ファイル1のデータカウント
    sheet_df1 = file1.parse(file1.sheet_names, header=None)
    sheet_names1 = file1.sheet_names

    for i, name in enumerate(sheet_names1):
        sheet_df1[i] = file1.parse(name)
        try :
            sig_ori = (sheet_df1[i]["volt"]).values
        except KeyError :
            print("not max data number")
            break
        sig_smooth = [(sig_ori[i]+sig_ori[i+1]+sig_ori[i+2])/3  for i in range(len(sig_ori)-2)]
        write_data(wb[name], sig_smooth, 2, 3, "smooth")

        sig_fire_a = [ 1 if(sig_smooth[i] - sig_smooth[i-1] <= 0) & (sig_smooth[i+1] - sig_smooth[i] >= 0) & (sig_ori[i+1] <= peak) \
                       else 0 for i in range(1, len(sig_smooth) -1)]

        write_data(wb[name], sig_fire_a, 3, 4, "spike_data")
        
        
#        sig_fire = [ 1 if(sig_smooth[i-3] > sig_smooth[i-2]) & (sig_smooth[i-2] > sig_smooth[i-1]) & (sig_smooth[i-1] > sig_smooth[i]) \
#                   & (sig_smooth[i] < sig_smooth[i+1]) & (sig_smooth[i+1] < sig_smooth[i+2]) & (sig_smooth[i+2] < sig_smooth[i+3]) else 0 for i in range(3, len(sig_smooth)-3)]

    
#read_path = "/st9/b009vb/01data"
read_path = sys.argv[1]
#save_path = ""
save_path = sys.argv[2]
#individual_path = "/restraint-25000/B39-HR/"
individual_path = sys.argv[3]
#file_name = "B39-restraint-R"
file_name = sys.argv[4]

#peak_file = openpyxl.load_workbook("peak_line.xlsx")
peak_file = openpyxl.load_workbook(sys.argv[5])
ws = peak_file.worksheets[0]

for row in ws.iter_rows(min_row=1):
    values=[]
    for col in row:
        values.append(col.value)
    if(values[0] == file_name) :
        peak = float(values[1])
    #print(values)

for i in range(2,6) :
    file1 = pd.ExcelFile(read_path + individual_path + file_name + str(i) + ".xlsx")
    wb = openpyxl.load_workbook(read_path + individual_path + file_name + str(i) + ".xlsx")
    
    inspect(wb, file1, peak)
    wb.save(file_name + str(i) + ".xlsx")
